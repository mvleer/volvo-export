"""
Layer 3 — Per-year output quality.

Validates every real volvo-trips-YYYY.xlsx file in volvo-trips/.
Parametrized over discovered files so a new year file automatically enters
coverage without any test changes.
"""

import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).parent.parent
OUTPUT_DIR = PROJECT_ROOT / "volvo-trips"

CANONICAL_HEADER = [
    "Category",
    "Started",
    "Start odometer (km)",
    "Start address",
    "Stopped",
    "End odometer (km)",
    "End address",
    "Duration",
    "Distance (km)",
    "Fuel consumption (l)",
    "Year",
    "l/100km",
    "Odo delta (km)",
]

ISO_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$")
YEAR_COL_PATTERN = re.compile(r"^\d{4}-([1-9]|1[0-2])$")  # no leading zero on month


def _real_year_files():
    return sorted(OUTPUT_DIR.glob("volvo-trips-[0-9][0-9][0-9][0-9].xlsx"))


def _read_xlsx_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:]]


# ---------------------------------------------------------------------------
# Structure guard
# ---------------------------------------------------------------------------

def test_output_files_exist():
    files = _real_year_files()
    assert len(files) > 0, (
        "No volvo-trips-YYYY.xlsx files found in volvo-trips/. "
        "Run: python volvo_trips_cleanup.py"
    )


# ---------------------------------------------------------------------------
# Per-file tests — parametrized over every volvo-trips-YYYY.xlsx
# ---------------------------------------------------------------------------

@pytest.fixture(params=_real_year_files(), ids=lambda p: p.name)
def year_file(request):
    return request.param


def test_header(year_file):
    wb = openpyxl.load_workbook(year_file)
    header = [cell.value for cell in next(wb.active.iter_rows(min_row=1, max_row=1))]
    assert header == CANONICAL_HEADER, (
        f"{year_file.name}: header mismatch\n  got:      {header}\n  expected: {CANONICAL_HEADER}"
    )


def test_sheet_name(year_file):
    wb = openpyxl.load_workbook(year_file)
    assert wb.active.title == "trips", (
        f"{year_file.name}: sheet name should be 'trips', got '{wb.active.title}'"
    )


def test_row_width(year_file):
    for i, row in enumerate(_read_xlsx_rows(year_file), 2):
        assert len(row) == len(CANONICAL_HEADER), (
            f"{year_file.name} row {i}: {len(row)} columns, expected {len(CANONICAL_HEADER)}"
        )


def test_iso_dates(year_file):
    for i, row in enumerate(_read_xlsx_rows(year_file), 2):
        for col in ("Started", "Stopped"):
            assert ISO_PATTERN.match(str(row[col])), (
                f"{year_file.name} row {i}: {col}='{row[col]}' is not ISO 8601 (YYYY-MM-DD HH:MM)"
            )


def test_no_unit_suffixes(year_file):
    for i, row in enumerate(_read_xlsx_rows(year_file), 2):
        for col in ("Distance (km)", "Fuel consumption (l)"):
            assert not re.search(r"[a-zA-Z]", str(row[col])), (
                f"{year_file.name} row {i}: {col}='{row[col]}' contains a unit suffix"
            )


def test_numeric_columns_are_numbers(year_file):
    """Distance and Fuel consumption must be stored as numbers, not strings."""
    for i, row in enumerate(_read_xlsx_rows(year_file), 2):
        for col in ("Distance (km)", "Fuel consumption (l)"):
            val = row[col]
            assert isinstance(val, (int, float)), (
                f"{year_file.name} row {i}: {col}='{val}' is not a number (type: {type(val).__name__})"
            )


def test_numeric_non_negative(year_file):
    for i, row in enumerate(_read_xlsx_rows(year_file), 2):
        for col in ("Distance (km)", "Fuel consumption (l)"):
            val = row[col]
            assert isinstance(val, (int, float)) and val >= 0, (
                f"{year_file.name} row {i}: {col}='{val}' must be a non-negative number"
            )


def test_distance_plausibility(year_file):
    """Sentinel for a missing ÷10 correction: no single trip should exceed 2000 km."""
    for i, row in enumerate(_read_xlsx_rows(year_file), 2):
        val = row["Distance (km)"]
        assert isinstance(val, (int, float)) and val <= 2000, (
            f"{year_file.name} row {i}: Distance='{val}' exceeds 2000 km "
            f"— ÷10 correction may be missing"
        )


def test_year_column_correctness(year_file):
    for i, row in enumerate(_read_xlsx_rows(year_file), 2):
        stopped = str(row["Stopped"])
        year_col = str(row["Year"])
        try:
            dt = datetime.strptime(stopped, "%Y-%m-%d %H:%M")
            expected = f"{dt.year}-{dt.month}"
            assert year_col == expected, (
                f"{year_file.name} row {i}: Year='{year_col}', "
                f"expected '{expected}' from Stopped='{stopped}'"
            )
        except ValueError:
            assert year_col == "", (
                f"{year_file.name} row {i}: Year should be empty for invalid Stopped date"
            )


def test_year_boundary(year_file):
    """Every Stopped date in the file must belong to the file's stated year."""
    file_year = int(year_file.stem.split("-")[-1])
    for i, row in enumerate(_read_xlsx_rows(year_file), 2):
        stopped = str(row["Stopped"])
        try:
            dt = datetime.strptime(stopped, "%Y-%m-%d %H:%M")
            assert dt.year == file_year, (
                f"{year_file.name} row {i}: Stopped='{stopped}' "
                f"belongs to {dt.year}, not {file_year}"
            )
        except ValueError:
            pass


# ---------------------------------------------------------------------------
# Category override round-trip tests
# ---------------------------------------------------------------------------

def test_category_override_preserved(tmp_path):
    """Non-Unassigned Category values in an existing XLSX survive a re-run."""
    from volvo_trips_cleanup import load_category_overrides, write_year_file, CANONICAL_HEADER

    rows = [
        {col: "" for col in CANONICAL_HEADER} | {
            "Category": "Maarten",
            "Started": "2025-01-10 08:00",
            "Start odometer (km)": "100000",
            "Stopped": "2025-01-10 08:30",
        }
    ]
    path = write_year_file(2025, rows, tmp_path)
    overrides = load_category_overrides(path)
    assert overrides == {("2025-01-10 08:00", "100000"): "Maarten"}


def test_unassigned_category_not_preserved(tmp_path):
    """Unassigned rows are not included in overrides."""
    from volvo_trips_cleanup import load_category_overrides, write_year_file, CANONICAL_HEADER

    rows = [
        {col: "" for col in CANONICAL_HEADER} | {
            "Category": "Unassigned",
            "Started": "2025-03-01 09:00",
            "Start odometer (km)": "101000",
            "Stopped": "2025-03-01 09:45",
        }
    ]
    path = write_year_file(2025, rows, tmp_path)
    overrides = load_category_overrides(path)
    assert overrides == {}


def test_category_override_int_odometer_roundtrip(tmp_path):
    """Category lookup must succeed when the odometer is an int (as clean_row produces).

    Regression test: run_pipeline was building lookup keys with a raw int odometer
    while load_category_overrides stored str(odo), so the dict lookup always missed.
    """
    from volvo_trips_cleanup import load_category_overrides, write_year_file, CANONICAL_HEADER

    odo = 115000  # int — exactly what clean_row writes into the row dict
    rows = [
        {col: "" for col in CANONICAL_HEADER} | {
            "Category": "Business",
            "Started": "2026-03-15 08:00",
            "Start odometer (km)": odo,
            "Stopped": "2026-03-15 08:30",
        }
    ]
    path = write_year_file(2026, rows, tmp_path)
    overrides = load_category_overrides(path)

    # Key as constructed in run_pipeline — both sides must be str
    key = (str(rows[0]["Started"]), str(odo))
    assert key in overrides, (
        f"Override not found — key type mismatch? "
        f"lookup={key!r}, stored={list(overrides)!r}"
    )
    assert overrides[key] == "Business"
