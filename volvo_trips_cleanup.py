import argparse
import csv
import re
from datetime import datetime
from pathlib import Path

import openpyxl

RAW_DIR = Path("raw")
OUTPUT_DIR = Path("volvo-trips")

ORIG_DATE_FORMAT = "%d/%m/%Y %H:%M"
ISO_DATE_FORMAT = "%Y-%m-%d %H:%M"

DATE_COLS = ["Started", "Stopped"]
REMOVE_COLS = {"Column1", "Title", "User Notes"}
DEDUP_KEY = ("Started", "Start odometer (km)")

CANONICAL_HEADER = [
    "Category",
    "Started",
    "Start odometer (km)",
    "Start address",
    "Stopped",
    "End odometer (km)",
    "End address",
    "Duration",
    "Distance (km)",
    "Fuel consumption (l)",
    "Year",
    "l/100km",
    "Odo delta (km)",
]

NUMERIC_COLS = {"Distance (km)", "Fuel consumption (l)", "Start odometer (km)", "End odometer (km)"}


def remove_unit(value: str) -> str:
    return re.sub(r"\s*[a-zA-Z]+$", "", value.strip())


def convert_date(value: str) -> str:
    try:
        return datetime.strptime(value.strip(), ORIG_DATE_FORMAT).strftime(ISO_DATE_FORMAT)
    except ValueError:
        return value


def clean_row(row: dict) -> dict:
    for col in DATE_COLS:
        if col in row:
            row[col] = convert_date(row[col])

    for col in ("Start odometer (km)", "End odometer (km)"):
        if col in row:
            try:
                row[col] = int(row[col])
            except (ValueError, TypeError):
                pass

    if "Distance (km)" in row:
        raw_val = row["Distance (km)"].strip()
        had_suffix = bool(re.search(r"[a-zA-Z]", raw_val))
        val = remove_unit(raw_val)
        try:
            normalized = val.replace(",", ".")
            numeric = float(normalized)
            # Divide by 10 when:
            #   - value had a unit suffix (legacy app style, e.g. "125,3 km"), OR
            #   - value is a pure integer (legacy app stores tenths-of-km as integers, e.g. "378")
            # Updated app format stores decimal km without suffix (e.g. "5.5") — no division.
            if had_suffix or "." not in normalized:
                numeric = round(numeric / 10, 2)
            else:
                numeric = round(numeric, 2)
            row["Distance (km)"] = numeric
        except ValueError:
            pass

    fuel_old = "Fuel consumption (litres)"
    fuel_new = "Fuel consumption (l)"
    if fuel_old in row:
        val = remove_unit(row.pop(fuel_old))
        try:
            row[fuel_new] = round(float(val.replace(",", ".")), 2)
        except ValueError:
            row[fuel_new] = val
    elif fuel_new in row:
        val = remove_unit(row[fuel_new])
        try:
            row[fuel_new] = round(float(val.replace(",", ".")), 2)
        except ValueError:
            row[fuel_new] = val

    stopped = row.get("Stopped", "")
    try:
        dt = datetime.strptime(stopped.strip(), ISO_DATE_FORMAT)
        row["Year"] = f"{dt.year}-{dt.month}"
    except ValueError:
        row["Year"] = ""

    try:
        start_odo = int(row.get("Start odometer (km)", ""))
        end_odo   = int(row.get("End odometer (km)", ""))
        row["Odo delta (km)"] = end_odo - start_odo
    except (TypeError, ValueError):
        row["Odo delta (km)"] = ""

    fuel     = row.get("Fuel consumption (l)")
    odo_delta = row.get("Odo delta (km)")
    try:
        if isinstance(fuel, (int, float)) and isinstance(odo_delta, (int, float)) and odo_delta > 0:
            row["l/100km"] = round(fuel / odo_delta * 100, 1)
        else:
            row["l/100km"] = ""
    except (TypeError, ZeroDivisionError):
        row["l/100km"] = ""

    return row


def read_raw_file(filepath: Path) -> list:
    rows = []
    with open(filepath, encoding="utf-8-sig") as f:
        sample = f.read(4096)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";"
        f.seek(0)
        reader = csv.DictReader(f, dialect=dialect)
        keep = set(
            h for h in (reader.fieldnames or [])
            if h and h.strip() and h not in REMOVE_COLS
        )
        for row in reader:
            cleaned = {k: v for k, v in row.items() if k in keep}
            rows.append(clean_row(cleaned))
    return rows


def read_raw_files(raw_dir: Path) -> list:
    rows = []
    for f in sorted(raw_dir.glob("*.csv")):
        rows.extend(read_raw_file(f))
    return rows


def merge_and_dedup(rows: list) -> list:
    seen = {}
    for row in rows:
        key = tuple(row.get(k, "") for k in DEDUP_KEY)
        if key not in seen:
            seen[key] = row
    return sorted(seen.values(), key=lambda r: r.get("Started", ""))


def split_by_year(rows: list) -> dict:
    by_year = {}
    for row in rows:
        try:
            year = datetime.strptime(row["Stopped"].strip(), ISO_DATE_FORMAT).year
        except (ValueError, KeyError):
            continue
        by_year.setdefault(year, []).append(row)
    return by_year


def load_category_overrides(path: Path) -> dict:
    """Read back manually assigned Category values from an existing year XLSX.
    Returns {(Started, Start odometer (km)): category} for non-Unassigned rows.
    """
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        cat_idx = headers.index("Category")
        started_idx = headers.index("Started")
        odo_idx = headers.index("Start odometer (km)")
    except ValueError:
        return {}
    overrides = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat = row[cat_idx]
        if cat and cat != "Unassigned":
            key = (str(row[started_idx]), str(row[odo_idx]))
            overrides[key] = cat
    return overrides


def write_year_file(year: int, rows: list, output_dir: Path) -> Path:
    outpath = output_dir / f"volvo-trips-{year}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "trips"
    ws.append(CANONICAL_HEADER)
    for row in rows:
        ws.append([row.get(col, "") for col in CANONICAL_HEADER])
    wb.save(outpath)
    return outpath


def run_pipeline(raw_dir: Path = RAW_DIR, output_dir: Path = OUTPUT_DIR) -> dict:
    raw_files = sorted(raw_dir.glob("*.csv"))
    if not raw_files:
        raise FileNotFoundError(f"No CSV files found in '{raw_dir}/'")

    output_dir.mkdir(parents=True, exist_ok=True)
    rows = read_raw_files(raw_dir)
    merged = merge_and_dedup(rows)
    by_year = split_by_year(merged)

    written = {}
    for year, year_rows in sorted(by_year.items()):
        outpath = output_dir / f"volvo-trips-{year}.xlsx"
        overrides = load_category_overrides(outpath)
        if overrides:
            for row in year_rows:
                key = (str(row.get("Started", "")), str(row.get("Start odometer (km)", "")))
                if key in overrides:
                    row["Category"] = overrides[key]
        path = write_year_file(year, year_rows, output_dir)
        written[year] = path
        print(f"Written {len(year_rows)} rows to '{path}' ({len(overrides)} category override(s) preserved).")

    return written


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Clean and split Volvo trip exports by year.")
    parser.add_argument("--raw-dir", type=Path, default=RAW_DIR,
                        help="Directory containing raw CSV exports (default: raw/)")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR,
                        help="Directory to write per-year output files (default: .)")
    args = parser.parse_args()
    run_pipeline(args.raw_dir, args.output_dir)
